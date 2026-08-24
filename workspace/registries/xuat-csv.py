#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Xuất 8 sổ đăng ký từ workbook Excel ra CSV đúng chuẩn.

    python xuat-csv.py so-dang-ky.xlsx registries

Vì sao cần script này thay vì "Save as CSV" trong Excel: Excel ghi kèm BOM và
xuống dòng CRLF. BOM dính vào tên cột đầu tiên nên `ma` bị đọc thành `﻿ma`,
và mọi mã trong sổ coi như rỗng — cả `project_check.py` của skill lẫn
`registry check` của BA Toolkit đều bắt lỗi này, nhưng bắt được sau khi BA đã
mất một buổi tự hỏi vì sao mã nào cũng báo sai.

Script ghi UTF-8 không BOM, xuống dòng LF, đúng thứ tự cột đã khai.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Thiếu openpyxl. Cài: python -m pip install openpyxl")

# Thứ tự cột lấy từ confluence_sync/registry.py — nguồn duy nhất.
COT = {
    "messages.csv": ["ma", "loai", "doi_tuong", "chu_de", "noi_dung", "thamso",
                     "dieu_kien_phat_sinh", "ghi_chu"],
    "usecases.csv": ["ma_uc", "stt", "ma_phan_he", "phan_he", "ten_uc",
                     "actor_chinh", "actor_phu", "bi_danh", "bmt", "so_trans",
                     "do_phuc_tap", "la_tu_dong", "pham_vi", "trang_thai",
                     "nhom_excel", "ghi_chu"],
    "roles.csv": ["ma", "ten", "tac_nhan_lien_quan", "pham_vi_mac_dinh",
                  "mo_ta", "ghi_chu"],
    "states.csv": ["ma", "doi_tuong", "ten", "ten_hien_thi", "mo_ta",
                   "thao_tac", "ghi_chu"],
    "participants.csv": ["ma", "ten", "loai", "mo_ta", "ghi_chu"],
    "components.csv": ["ma", "ten", "loai", "mo_ta", "ghi_chu"],
    "objects.csv": ["ma", "ten", "ten_hien_thi", "mo_ta", "ghi_chu"],
    "groups.csv": ["ma", "ten", "ghi_chu"],
}

# Mã của các dòng ví dụ shipped kèm workbook. Còn sót là BA quên xoá.
VI_DU = {
    "INF_001", "SUC_001", "MAIL_001", "UC-0301", "UC-0302", "ROLE-QTHT",
    "ROLE-NVNV", "ST-NGUOIDUNG-01", "ST-NGUOIDUNG-02", "PAR-CIC", "PAR-CBTD",
    "CMP_DANHMUC", "CMP_TIMKIEM", "NGUOIDUNG", "SANPHAM", "GRP-QLNSD-01",
    "GRP-QLSP-01",
}

DUNG_LAI = "◆"          # dòng mở đầu khối chú thích cuối sheet


def _txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def doc_sheet(ws, cols: list[str]) -> tuple[list[list[str]], list[str]]:
    """Đọc tới dòng trống đầu tiên. Trả về (dòng dữ liệu, cảnh báo)."""
    canh_bao: list[str] = []
    head = [_txt(c.value) for c in ws[1]][:len(cols)]
    if head != cols:
        canh_bao.append(
            f"tiêu đề lệch đặc tả.\n      trên sheet : {head}\n      cần có     : {cols}")
        return [], canh_bao

    rows: list[list[str]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        cells = [_txt(v) for v in (r or ())][:len(cols)]
        cells += [""] * (len(cols) - len(cells))
        if cells and cells[0].startswith(DUNG_LAI):
            break
        if not any(cells):
            break
        rows.append(cells)
    return rows, canh_bao


def main() -> int:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    src, dst = Path(sys.argv[1]), Path(sys.argv[2])
    if not src.is_file():
        print(f"LỖI: không thấy {src}")
        return 1
    dst.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(src, data_only=True)
    tong, con_vi_du, loi = 0, [], False

    for fname, cols in COT.items():
        ten = fname.replace(".csv", "")
        if ten not in wb.sheetnames:
            print(f"  [LỖI ] thiếu sheet `{ten}`")
            loi = True
            continue
        rows, canh_bao = doc_sheet(wb[ten], cols)
        for c in canh_bao:
            print(f"  [LỖI ] {fname}: {c}")
            loi = True
        if canh_bao:
            continue

        khoa = [r[0] for r in rows]
        trung = sorted({k for k in khoa if khoa.count(k) > 1 and k})
        if trung:
            print(f"  [LỖI ] {fname}: mã trùng — {', '.join(trung)}")
            loi = True
        rong = sum(1 for k in khoa if not k)
        if rong:
            print(f"  [LỖI ] {fname}: {rong} dòng bỏ trống cột khoá `{cols[0]}`")
            loi = True
        con_vi_du += [f"{fname}:{k}" for k in khoa if k in VI_DU]

        with (dst / fname).open("w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh, lineterminator="\n")
            w.writerow(cols)
            w.writerows(rows)
        tong += len(rows)
        print(f"  [  ok ] {fname:18} {len(rows):4} dòng")

    if con_vi_du:
        print(f"\n  [CẢNH] còn {len(con_vi_du)} dòng ví dụ mẫu chưa xoá: "
              f"{', '.join(con_vi_du[:6])}"
              + (" …" if len(con_vi_du) > 6 else ""))
        print("        Ví dụ mẫu publish lên Confluence là dữ liệu rác dùng chung.")

    print(f"\n  {tong} dòng ra {dst}")
    if loi:
        print("  → CÓ LỖI. Sửa trên workbook rồi chạy lại.")
        return 1
    print("  → Xong. Kiểm tiếp bằng project_check.py hoặc `registry check`.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
